from pymongo import MongoClient # エクセルファイルに書き込み用
import openpyxl
import score
client = MongoClient('localhost', 27017)
db = client.score_database
collection = db.test_collection
file_name = 'sakaguti.xlsx'
sheet_name = '2019年OSIデータ'
file = openpyxl.load_workbook(file_name)
sheet = file[sheet_name]

colm = 3
for i in range(0,14):
    number = score.number[i]
    for a in range(2,63):
        value = sheet.cell(row=a, column=colm).value
        print(value)
        score_value = collection.find_one({'$and': [{'number':number},{'origin':value}]})
        try:
            print(score_value)
            sheet.cell(row=a, column = colm).value = score_value['score']
        except:
            print('continue')
            continue
    colm += 1

file.save('sakaguti_answer.xlsx')


